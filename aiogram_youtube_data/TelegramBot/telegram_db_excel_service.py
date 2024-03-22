import sqlite3
import io
from openpyxl import Workbook


def send_excel_file() -> io.BytesIO:
    # Создаем байтовый поток для хранения данных Excel
    excel_buffer = io.BytesIO()

    # Подключаемся к базе данных SQLite
    conn = sqlite3.connect('db.sqlite3')
    cursor = conn.cursor()

    try:
        # Создаем новый файл Excel
        workbook = Workbook()

        # Получаем данные из таблицы video_info
        cursor.execute("SELECT * FROM video_info")
        data_video_info = cursor.fetchall()

        # Создаем лист для данных из таблицы video_info
        video_info_sheet = workbook.create_sheet(title="Video Info")

        # Записываем заголовки столбцов для таблицы video_info
        headers_video_info = [description[0] for description in cursor.description]
        for col, header in enumerate(headers_video_info, start=1):
            video_info_sheet.cell(row=1, column=col, value=header)

        # Записываем данные в лист для таблицы video_info
        for row, row_data in enumerate(data_video_info, start=2):
            for col, value in enumerate(row_data, start=1):
                video_info_sheet.cell(row=row, column=col, value=value)

        # Получаем данные из таблицы playlist_info
        cursor.execute("SELECT * FROM playlist_info")
        data_playlist_info = cursor.fetchall()

        # Создаем лист для данных из таблицы playlist_info
        playlist_info_sheet = workbook.create_sheet(title="Playlist Info")

        # Записываем заголовки столбцов для таблицы playlist_info
        headers_playlist_info = [description[0] for description in cursor.description]
        for col, header in enumerate(headers_playlist_info, start=1):
            playlist_info_sheet.cell(row=1, column=col, value=header)

        # Записываем данные в лист для таблицы playlist_info
        for row, row_data in enumerate(data_playlist_info, start=2):
            for col, value in enumerate(row_data, start=1):
                playlist_info_sheet.cell(row=row, column=col, value=value)

        # Получаем данные из таблицы channel_info
        cursor.execute("SELECT * FROM channel_info")
        data_channel_info = cursor.fetchall()

        # Создаем лист для данных из таблицы channel_info
        channel_info_sheet = workbook.create_sheet(title="Channel Info")

        # Записываем заголовки столбцов для таблицы channel_info
        headers_channel_info = [description[0] for description in cursor.description]
        for col, header in enumerate(headers_channel_info, start=1):
            channel_info_sheet.cell(row=1, column=col, value=header)

        # Записываем данные в лист для таблицы channel_info
        for row, row_data in enumerate(data_channel_info, start=2):
            for col, value in enumerate(row_data, start=1):
                channel_info_sheet.cell(row=row, column=col, value=value)

        # Удаляем дефолтный лист, если нужно
        workbook.remove(workbook["Sheet"])

        # Сохраняем файл Excel в байтовый поток
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
    finally:
        # Закрываем соединение с базой данных
        conn.close()

    return excel_buffer
